#!/usr/bin/env python3
"""
Create Excel Forecast Template for Call Center Workforce Planning

This script generates a comprehensive Excel workbook with pre-built worksheets
for forecasting call center volumes using multiple methods.

Requires: openpyxl (pip install openpyxl)
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("ERROR: openpyxl library not found.")
    print("Please install it using: pip install openpyxl")
    exit(1)

def create_instructions_sheet(wb):
    """Create the Instructions worksheet"""
    ws = wb.create_sheet("📖 Instructions", 0)

    # Title
    ws['A1'] = "Call Center Forecast Template - User Guide"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws.merge_cells('A1:E1')

    # Content
    instructions = [
        ["", ""],
        ["📋 Quick Start (5 minutes)", ""],
        ["1. Go to 'Data Input' sheet", ""],
        ["2. Paste your historical data (Date, Time, Calls_Offered, AHT)", ""],
        ["3. Go to 'FORECAST.ETS' sheet", ""],
        ["4. Review the automatically generated forecast", ""],
        ["5. Check 'Accuracy Dashboard' for MAPE score", ""],
        ["", ""],
        ["📊 Worksheet Overview", ""],
        ["", ""],
        ["Sheet Name", "Purpose"],
        ["Data Input", "Paste your historical call data here (min 12 months)"],
        ["FORECAST.ETS", "Automated forecasting using Excel 2016+ built-in function"],
        ["Seasonal Decomposition", "Manual method - works in any Excel version"],
        ["Simple Exp Smoothing", "Quick short-term (1-3 days) forecast"],
        ["Accuracy Dashboard", "Track forecast accuracy (MAPE, MAE, RMSE)"],
        ["Event Calendar", "List special events and volume impacts"],
        ["Staffing Calculator", "Convert forecasted calls to required agents"],
        ["", ""],
        ["💡 Tips for Best Results", ""],
        ["• Use at least 12 months of data (24-36 months recommended)", ""],
        ["• Data should be in 15-30 minute intervals", ""],
        ["• Remove system outages and data errors before forecasting", ""],
        ["• Update forecasts weekly with latest data", ""],
        ["• Track accuracy and adjust methods if MAPE > 10%", ""],
        ["", ""],
        ["📈 Accuracy Targets", ""],
        ["MAPE < 5%  = Excellent (highly reliable)", ""],
        ["MAPE 5-7%  = Good (suitable for most decisions)", ""],
        ["MAPE 7-10% = Acceptable (add safety buffers)", ""],
        ["MAPE > 10% = Poor (need more data or different method)", ""],
        ["", ""],
        ["🔗 Resources", ""],
        ["• Quick Start Guide: FORECASTING_QUICKSTART.md", ""],
        ["• Comprehensive Guide: call_center_forecasting_guide.md", ""],
        ["• GitHub: https://github.com/dpereda/call-center-workforce-planning", ""],
    ]

    for idx, row in enumerate(instructions, start=2):
        ws[f'A{idx}'] = row[0]
        if idx in [3, 10, 21, 28, 34]:  # Headers
            ws[f'A{idx}'].font = Font(size=12, bold=True, color="1F4E78")
        if idx == 11:  # Table header
            ws[f'A{idx}'].font = Font(bold=True)
            ws[f'B{idx}'].font = Font(bold=True)
            ws[f'A{idx}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            ws[f'B{idx}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        if len(row) > 1 and row[1]:
            ws[f'B{idx}'] = row[1]

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 60

    return ws

def create_data_input_sheet(wb):
    """Create the Data Input worksheet"""
    ws = wb.create_sheet("📥 Data Input")

    # Title
    ws['A1'] = "Historical Call Data"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws.merge_cells('A1:F1')

    # Instructions
    ws['A2'] = "Paste your historical data below (minimum 12 months recommended)"
    ws['A2'].font = Font(italic=True, color="7F7F7F")
    ws.merge_cells('A2:F2')

    # Headers
    headers = ["Date", "Time_Interval", "Calls_Offered", "AHT_Seconds", "Calls_Answered", "Calls_Abandoned"]
    header_row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Sample data
    sample_data = [
        ["1/1/25", "08:00-08:15", 16, 270, 15, 1],
        ["1/1/25", "08:15-08:30", 12, 268, 12, 0],
        ["1/1/25", "08:30-08:45", 9, 275, 9, 0],
        ["(Paste your data starting from row 5)", "", "", "", "", ""],
    ]

    for row_idx, row_data in enumerate(sample_data, start=5):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

    return ws

def create_forecast_ets_sheet(wb):
    """Create the FORECAST.ETS worksheet"""
    ws = wb.create_sheet("📈 FORECAST.ETS")

    # Title
    ws['A1'] = "FORECAST.ETS - Automated Forecasting"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws.merge_cells('A1:G1')

    # Instructions
    ws['A2'] = "This method uses Excel 2016+ Triple Exponential Smoothing (Holt-Winters). Requires Excel 2016 or newer."
    ws['A2'].font = Font(italic=True, color="7F7F7F")
    ws.merge_cells('A2:G2')

    # Setup section
    ws['A4'] = "📊 Setup"
    ws['A4'].font = Font(size=12, bold=True, color="1F4E78")

    ws['A5'] = "Historical Data Range:"
    ws['A6'] = "Forecast Start Date:"
    ws['A7'] = "Forecast Periods:"
    ws['A8'] = "Seasonality (96 = daily for 15-min):"

    ws['B5'] = "='📥 Data Input'!A5:A1000"
    ws['B6'] = "=TODAY()+1"
    ws['B7'] = 672
    ws['C7'] = "(672 = 1 week of 15-min intervals)"
    ws['C7'].font = Font(italic=True, size=9, color="7F7F7F")
    ws['B8'] = 96

    # Headers
    headers = ["Date", "Time_Interval", "Forecasted_Calls", "Lower_Bound_95%", "Upper_Bound_95%", "Event_Adjustment", "Final_Forecast"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=10, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Formula examples (row 11)
    ws['A11'] = "=$B$6"
    ws['B11'] = '="08:00-08:15"'
    ws['C11'] = '=IFERROR(FORECAST.ETS($A11, \'📥 Data Input\'!$C$5:$C$1000, \'📥 Data Input\'!$A$5:$A$1000, $B$8, 1, 1), "Need Excel 2016+")'
    ws['D11'] = '=IFERROR(C11 - FORECAST.ETS.CONFINT($A11, \'📥 Data Input\'!$C$5:$C$1000, \'📥 Data Input\'!$A$5:$A$1000, 0.95, $B$8), "")'
    ws['E11'] = '=IFERROR(C11 + FORECAST.ETS.CONFINT($A11, \'📥 Data Input\'!$C$5:$C$1000, \'📥 Data Input\'!$A$5:$A$1000, 0.95, $B$8), "")'
    ws['F11'] = 1.0
    ws['G11'] = "=C11*F11"

    # Instructions for copying
    ws['A13'] = "Instructions: Copy formulas in row 11 down for all forecast periods"
    ws['A13'].font = Font(italic=True, color="7F7F7F")
    ws.merge_cells('A13:G13')

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18

    return ws

def create_seasonal_decomp_sheet(wb):
    """Create the Seasonal Decomposition worksheet"""
    ws = wb.create_sheet("📉 Seasonal Decomp")

    # Title
    ws['A1'] = "Seasonal Decomposition Method"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws.merge_cells('A1:F1')

    # Instructions
    ws['A2'] = "This method works in ANY Excel version. Formula: Forecast = Trend × Seasonal_Index"
    ws['A2'].font = Font(italic=True, color="7F7F7F")
    ws.merge_cells('A2:F2')

    # Step 1: Calculate averages
    ws['A4'] = "Step 1: Calculate Seasonal Indices"
    ws['A4'].font = Font(size=12, bold=True, color="1F4E78")

    ws['A5'] = "Time Interval"
    ws['B5'] = "Avg Calls (Historical)"
    ws['C5'] = "Overall Average"
    ws['D5'] = "Seasonal Index"

    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}5'].font = Font(bold=True)
        ws[f'{col}5'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # Sample seasonal index calculations
    time_intervals = ["08:00-08:15", "08:15-08:30", "08:30-08:45", "...(continue for all 96 intervals)"]
    for idx, interval in enumerate(time_intervals, start=6):
        ws[f'A{idx}'] = interval
        if idx == 6:
            ws[f'B{idx}'] = '=AVERAGEIF(\'📥 Data Input\'!$B$5:$B$1000, A6, \'📥 Data Input\'!$C$5:$C$1000)'
            ws[f'C{idx}'] = '=AVERAGE(\'📥 Data Input\'!$C$5:$C$1000)'
            ws[f'D{idx}'] = '=B6/C6'

    # Step 2: Calculate trend
    ws['A12'] = "Step 2: Calculate Trend"
    ws['A12'].font = Font(size=12, bold=True, color="1F4E78")

    ws['A13'] = "Use linear regression or moving average to determine growth trend"
    ws['A13'].font = Font(italic=True)
    ws['B13'] = "Trend_Multiplier:"
    ws['C13'] = 1.07
    ws['D13'] = "(Example: 1.07 = 7% annual growth)"
    ws['D13'].font = Font(italic=True, size=9, color="7F7F7F")

    # Step 3: Generate forecast
    ws['A16'] = "Step 3: Generate Forecast"
    ws['A16'].font = Font(size=12, bold=True, color="1F4E78")

    headers = ["Date", "Time_Interval", "Base_Calls", "Seasonal_Index", "Trend", "Forecast"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=17, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    ws['A18'] = "=TODAY()+1"
    ws['B18'] = "08:00-08:15"
    ws['C18'] = "=AVERAGE('📥 Data Input'!$C$5:$C$1000)"
    ws['D18'] = "=VLOOKUP(B18, $A$6:$D$101, 4, FALSE)"
    ws['E18'] = "=$C$13"
    ws['F18'] = "=C18*D18*E18"

    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15

    return ws

def create_exponential_smoothing_sheet(wb):
    """Create the Simple Exponential Smoothing worksheet"""
    ws = wb.create_sheet("⚡ Simple Exp Smooth")

    # Title
    ws['A1'] = "Simple Exponential Smoothing"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws.merge_cells('A1:E1')

    # Instructions
    ws['A2'] = "Best for short-term forecasts (1-3 days). Formula: Forecast = α × Actual + (1-α) × Previous_Forecast"
    ws['A2'].font = Font(italic=True, color="7F7F7F")
    ws.merge_cells('A2:E2')

    # Setup
    ws['A4'] = "Alpha (α):"
    ws['B4'] = 0.3
    ws['C4'] = "(0.1-0.3 for stable, 0.4-0.6 for volatile)"
    ws['C4'].font = Font(italic=True, size=9, color="7F7F7F")

    ws['A5'] = "Initial Forecast:"
    ws['B5'] = '=AVERAGE(\'📥 Data Input\'!$C$5:$C$100)'

    # Headers
    headers = ["Date", "Time_Interval", "Actual_Calls", "Forecast", "Error"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # Sample formulas
    ws['A8'] = "='📥 Data Input'!A5"
    ws['B8'] = "='📥 Data Input'!B5"
    ws['C8'] = "='📥 Data Input'!C5"
    ws['D8'] = "=$B$5"
    ws['E8'] = "=C8-D8"

    ws['A9'] = "='📥 Data Input'!A6"
    ws['B9'] = "='📥 Data Input'!B6"
    ws['C9'] = "='📥 Data Input'!C6"
    ws['D9'] = "=$B$4*C8+(1-$B$4)*D8"
    ws['E9'] = "=C9-D9"

    ws['A11'] = "Copy row 9 formulas down for all historical data, then continue for forecast periods"
    ws['A11'].font = Font(italic=True, color="7F7F7F")
    ws.merge_cells('A11:E11')

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12

    return ws

def create_accuracy_dashboard(wb):
    """Create the Accuracy Dashboard worksheet"""
    ws = wb.create_sheet("📊 Accuracy Dashboard")

    # Title
    ws['A1'] = "Forecast Accuracy Metrics"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws.merge_cells('A1:D1')

    # Key Metrics
    ws['A3'] = "📈 Key Metrics"
    ws['A3'].font = Font(size=12, bold=True, color="1F4E78")

    metrics = [
        ["Metric", "Formula", "Your Result", "Target"],
        ["MAPE (%)", "=AVERAGE(ABS((Actual-Forecast)/Actual)*100)", "", "< 7%"],
        ["MAE", "=AVERAGE(ABS(Actual-Forecast))", "", "< 5 calls"],
        ["RMSE", "=SQRT(AVERAGE((Actual-Forecast)^2))", "", "< 8 calls"],
        ["Bias", "=AVERAGE(Forecast-Actual)", "", "~0 (unbiased)"],
    ]

    for row_idx, row in enumerate(metrics, start=4):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 4:  # Header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # Instructions
    ws['A10'] = "📝 How to Use This Dashboard"
    ws['A10'].font = Font(size=12, bold=True, color="1F4E78")

    ws['A11'] = "1. Replace 'Actual' and 'Forecast' in formulas above with your data ranges"
    ws['A12'] = "2. MAPE < 7% is good for call center forecasting"
    ws['A13'] = "3. Track these metrics weekly to monitor forecast quality"
    ws['A14'] = "4. If MAPE > 10%, review your method or data quality"

    # Detail section
    ws['A16'] = "📋 Detailed Error Analysis"
    ws['A16'].font = Font(size=12, bold=True, color="1F4E78")

    headers = ["Date", "Time", "Actual", "Forecast", "Error", "Abs_Error", "Pct_Error", "Sq_Error"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=17, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # Sample formulas
    ws['A18'] = "='📥 Data Input'!A5"
    ws['B18'] = "='📥 Data Input'!B5"
    ws['C18'] = "='📥 Data Input'!C5"
    ws['D18'] = "='📈 FORECAST.ETS'!G11"
    ws['E18'] = "=D18-C18"
    ws['F18'] = "=ABS(E18)"
    ws['G18'] = "=IF(C18<>0, ABS(E18/C18)*100, 0)"
    ws['H18'] = "=E18^2"

    # Column widths
    for col in ['A', 'B']:
        ws.column_dimensions[col].width = 12
    for col in ['C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 13

    return ws

def create_event_calendar(wb):
    """Create the Event Calendar worksheet"""
    ws = wb.create_sheet("📅 Event Calendar")

    # Title
    ws['A1'] = "Special Events Calendar"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws.merge_cells('A1:E1')

    # Instructions
    ws['A2'] = "List special events that impact call volumes. Use multipliers in forecast sheets."
    ws['A2'].font = Font(italic=True, color="7F7F7F")
    ws.merge_cells('A2:E2')

    # Headers
    headers = ["Date", "Event_Name", "Event_Type", "Expected_Impact", "Notes"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # Sample events
    events = [
        ["2/14/25", "Valentine's Day Promo", "Marketing Campaign", "1.25 (+25%)", "Email blast + social media"],
        ["3/15/25", "Spring Product Launch", "Product Launch", "1.40 (+40%)", "New product line announcement"],
        ["4/1/25", "Monthly Billing", "Billing Cycle", "1.15 (+15%)", "Standard monthly spike"],
        ["7/4/25", "Independence Day", "Holiday (Closed)", "0.00 (-100%)", "Office closed"],
        ["11/24/25", "Black Friday", "Peak Event", "1.60 (+60%)", "Major sales event"],
    ]

    for row_idx, event in enumerate(events, start=5):
        for col_idx, value in enumerate(event, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Reference guide
    ws['A12'] = "📚 Typical Impact Multipliers"
    ws['A12'].font = Font(size=11, bold=True, color="1F4E78")

    impact_guide = [
        ["Event Type", "Typical Multiplier"],
        ["Email Campaign", "1.15 - 1.25 (+15-25%)"],
        ["Product Launch", "1.30 - 1.50 (+30-50%)"],
        ["Billing Cycle", "1.10 - 1.15 (+10-15%)"],
        ["Holiday (before)", "1.20 - 1.40 (+20-40%)"],
        ["Holiday (closed)", "0.00 (-100%)"],
        ["Black Friday/Peak", "1.50 - 1.80 (+50-80%)"],
    ]

    for row_idx, row in enumerate(impact_guide, start=13):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 13:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 35

    return ws

def create_staffing_calculator(wb):
    """Create the Staffing Calculator worksheet"""
    ws = wb.create_sheet("👥 Staffing Calculator")

    # Title
    ws['A1'] = "Staffing Requirements Calculator"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws.merge_cells('A1:H1')

    # Instructions
    ws['A2'] = "Convert forecasted call volumes to required agent staffing using Square Root Staffing Rule"
    ws['A2'].font = Font(italic=True, color="7F7F7F")
    ws.merge_cells('A2:H2')

    # Setup
    ws['A4'] = "⚙️ Configuration"
    ws['A4'].font = Font(size=11, bold=True, color="1F4E78")

    ws['A5'] = "Interval Length (seconds):"
    ws['B5'] = 900
    ws['C5'] = "(900 = 15 minutes)"
    ws['C5'].font = Font(italic=True, size=9, color="7F7F7F")

    ws['A6'] = "Expected AHT (seconds):"
    ws['B6'] = 270

    ws['A7'] = "K-value (safety factor):"
    ws['B7'] = 1.6
    ws['C7'] = "(1.4-2.0 range, higher = better service)"
    ws['C7'].font = Font(italic=True, size=9, color="7F7F7F")

    ws['A8'] = "Service Level Target:"
    ws['B8'] = "80/90"

    # Headers
    headers = ["Date", "Time", "Forecasted_Calls", "Traffic (Erlangs)", "Base_Agents", "Safety_Buffer", "Required_Agents", "Notes"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=10, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Sample formulas
    ws['A11'] = "='📈 FORECAST.ETS'!A11"
    ws['B11'] = "='📈 FORECAST.ETS'!B11"
    ws['C11'] = "='📈 FORECAST.ETS'!G11"
    ws['D11'] = "=(C11*$B$6)/$B$5"
    ws['E11'] = "=D11"
    ws['F11'] = "=$B$7*SQRT(D11)"
    ws['G11'] = "=CEILING(E11+F11, 1)"
    ws['H11'] = '=IF(G11>20,"Peak Period",IF(G11<5,"Low Volume","Normal"))'

    ws['A13'] = "Copy formulas down for all forecast periods"
    ws['A13'].font = Font(italic=True, color="7F7F7F")
    ws.merge_cells('A13:H13')

    # Summary section
    ws['A15'] = "📊 Summary Statistics"
    ws['A15'].font = Font(size=11, bold=True, color="1F4E78")

    summary = [
        ["Metric", "Formula", "Result"],
        ["Average Agents Required", "=AVERAGE(G11:G107)", ""],
        ["Peak Agents Required", "=MAX(G11:G107)", ""],
        ["Minimum Agents Required", "=MIN(G11:G107)", ""],
        ["Total Agent Hours (Weekly)", "=SUM(G11:G107)/4", ""],
    ]

    for row_idx, row in enumerate(summary, start=16):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 16:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['D'].width = 17
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 17
    ws.column_dimensions['H'].width = 15

    return ws

def main():
    """Main function to create the Excel workbook"""
    print("Creating Call Center Forecast Template...")

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create all worksheets
    print("  ✓ Creating Instructions sheet")
    create_instructions_sheet(wb)

    print("  ✓ Creating Data Input sheet")
    create_data_input_sheet(wb)

    print("  ✓ Creating FORECAST.ETS sheet")
    create_forecast_ets_sheet(wb)

    print("  ✓ Creating Seasonal Decomposition sheet")
    create_seasonal_decomp_sheet(wb)

    print("  ✓ Creating Simple Exponential Smoothing sheet")
    create_exponential_smoothing_sheet(wb)

    print("  ✓ Creating Accuracy Dashboard")
    create_accuracy_dashboard(wb)

    print("  ✓ Creating Event Calendar")
    create_event_calendar(wb)

    print("  ✓ Creating Staffing Calculator")
    create_staffing_calculator(wb)

    # Save workbook
    filename = "call_center_forecast_template.xlsx"
    wb.save(filename)

    print(f"\n✅ Successfully created {filename}")
    print(f"\n📋 Template includes:")
    print(f"   • Instructions and user guide")
    print(f"   • Data input worksheet with validation")
    print(f"   • FORECAST.ETS (Excel 2016+)")
    print(f"   • Seasonal Decomposition (any Excel version)")
    print(f"   • Simple Exponential Smoothing")
    print(f"   • Accuracy tracking dashboard (MAPE, MAE, RMSE)")
    print(f"   • Event calendar for special events")
    print(f"   • Staffing calculator (Square Root method)")
    print(f"\n🎯 Next steps:")
    print(f"   1. Open {filename} in Excel")
    print(f"   2. Paste your historical data in 'Data Input' sheet")
    print(f"   3. Review forecasts in 'FORECAST.ETS' sheet")
    print(f"   4. Check accuracy in 'Accuracy Dashboard'")

if __name__ == '__main__':
    main()
