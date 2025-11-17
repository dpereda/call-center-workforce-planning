#!/usr/bin/env python3
"""
Service Level Calculator - Excel Worksheet Generator

This script creates a new worksheet in erlang_c_staffing_forecast.xlsx that allows users to:
1. Input their agent schedule
2. Apply a shrinkage rate
3. Calculate achievable service levels using Erlang C formulas

The worksheet calculates:
- Net agents (after shrinkage)
- Service level achievement (80/90 target)
- Average Speed of Answer (ASA)
- Occupancy rate
- Staffing gap analysis
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
import pandas as pd
from datetime import datetime, timedelta

def create_service_level_worksheet():
    """
    Create a new worksheet for service level calculations based on agent schedules
    """

    # Load the existing workbook
    try:
        workbook = openpyxl.load_workbook('erlang_c_staffing_forecast.xlsx')
        print("✓ Loaded existing workbook: erlang_c_staffing_forecast.xlsx")
    except FileNotFoundError:
        print("! Workbook not found. Creating new workbook.")
        workbook = openpyxl.Workbook()
        # Remove default sheet
        if 'Sheet' in workbook.sheetnames:
            del workbook['Sheet']

    # Create new worksheet
    if 'Schedule_Service_Level' in workbook.sheetnames:
        print("! Worksheet 'Schedule_Service_Level' already exists. Removing old version.")
        del workbook['Schedule_Service_Level']

    ws = workbook.create_sheet('Schedule_Service_Level', 0)
    print("✓ Created new worksheet: Schedule_Service_Level")

    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    subheader_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
    subheader_font = Font(bold=True, size=10)
    input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    calc_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 16
    ws.column_dimensions['J'].width = 18
    ws.column_dimensions['K'].width = 20
    ws.column_dimensions['L'].width = 14
    ws.column_dimensions['M'].width = 14
    ws.column_dimensions['N'].width = 16

    # ===== TITLE AND INSTRUCTIONS =====
    ws.merge_cells('A1:N1')
    ws['A1'] = 'SERVICE LEVEL CALCULATOR - Agent Schedule → Achievable Service Levels'
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    ws['A1'].alignment = center_align

    ws.merge_cells('A2:N2')
    ws['A2'] = 'Instructions: Enter your Shrinkage Rate in cell E4. Enter your Scheduled Agents in column D (starting row 7). The calculator will show achievable service levels.'
    ws['A2'].font = Font(italic=True, size=9)
    ws['A2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30

    # ===== SHRINKAGE INPUT SECTION =====
    ws['A4'] = 'Shrinkage Rate:'
    ws['A4'].font = Font(bold=True, size=11)
    ws['E4'] = 0.25  # Default 25% shrinkage
    ws['E4'].number_format = '0%'
    ws['E4'].fill = input_fill
    ws['E4'].font = Font(bold=True, size=12, color="C00000")
    ws['E4'].border = border

    ws.merge_cells('F4:N4')
    ws['F4'] = 'Typical shrinkage: 25-30% (includes breaks, lunch, meetings, training, absenteeism)'
    ws['F4'].font = Font(italic=True, size=9, color="7F7F7F")

    # ===== COLUMN HEADERS =====
    row = 6

    # Section labels
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'] = 'SCHEDULE INPUTS'
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].alignment = center_align

    ws.merge_cells(f'D{row}:E{row}')
    ws[f'D{row}'] = 'STAFFING'
    ws[f'D{row}'].fill = header_fill
    ws[f'D{row}'].font = header_font
    ws[f'D{row}'].alignment = center_align

    ws.merge_cells(f'F{row}:I{row}')
    ws[f'F{row}'] = 'CALCULATIONS'
    ws[f'F{row}'].fill = header_fill
    ws[f'F{row}'].font = header_font
    ws[f'F{row}'].alignment = center_align

    ws.merge_cells(f'J{row}:N{row}')
    ws[f'J{row}'] = 'OUTPUTS & METRICS'
    ws[f'J{row}'].fill = header_fill
    ws[f'J{row}'].font = header_font
    ws[f'J{row}'].alignment = center_align

    # Column headers
    row = 7
    headers = [
        'Day', 'Date', 'Time Interval',
        'Scheduled Agents', 'Net Agents',
        'Calls Offered', 'AHT (sec)', 'Traffic (Erlangs)', 'Req. Agents',
        'Erlang C Prob.', 'Service Level %', 'ASA (sec)', 'Occupancy %', 'Staffing Gap'
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.fill = subheader_fill
        cell.font = subheader_font
        cell.alignment = center_align
        cell.border = border

    # ===== LOAD FORECAST DATA =====
    try:
        forecast_df = pd.read_csv('erlang_c_staffing_forecast.csv')
        print(f"✓ Loaded forecast data: {len(forecast_df)} rows")
    except FileNotFoundError:
        print("! Forecast file not found. Using sample data.")
        # Create sample data
        forecast_df = pd.DataFrame({
            'Day': ['Monday'] * 36,
            'Date': ['2024-01-08'] * 36,
            'Time_Interval': [f"{h:02d}:{m:02d}-{h if m < 45 else h+1:02d}:{m+15 if m < 45 else (m+15)%60:02d}"
                             for h in range(8, 17) for m in range(0, 60, 15)],
            'Calls_Offered': [18, 22, 25, 28, 30, 32, 35, 33, 30, 28, 32, 34,
                            30, 28, 30, 32, 28, 25, 22, 20, 18, 16, 15, 14,
                            12, 11, 10, 9, 8, 7, 8, 9, 10, 11, 12, 13],
            'Average_Handle_Time_Seconds': [260] * 36,
            'Required_Agents': [6, 7, 8, 9, 9, 10, 11, 10, 9, 9, 10, 10,
                              9, 9, 9, 10, 9, 8, 7, 7, 6, 6, 5, 5,
                              5, 4, 4, 4, 3, 3, 3, 4, 4, 4, 5, 5]
        })

    # ===== POPULATE DATA ROWS =====
    data_start_row = 8

    for idx, row_data in forecast_df.head(36).iterrows():  # One day sample
        row_num = data_start_row + idx

        # Input columns (from forecast)
        ws[f'A{row_num}'] = row_data.get('Day', 'Monday')
        ws[f'B{row_num}'] = row_data.get('Date', '2024-01-08')
        ws[f'C{row_num}'] = row_data.get('Time_Interval', f'{8+idx//4:02d}:00')

        # Scheduled Agents (USER INPUT - leave empty for user to fill)
        ws[f'D{row_num}'].fill = input_fill
        ws[f'D{row_num}'].border = border
        ws[f'D{row_num}'].alignment = center_align
        ws[f'D{row_num}'].value = row_data.get('Required_Agents', 10)  # Placeholder

        # Net Agents formula: Scheduled × (1 - Shrinkage)
        ws[f'E{row_num}'] = f'=ROUND(D{row_num}*(1-$E$4),1)'
        ws[f'E{row_num}'].number_format = '0.0'
        ws[f'E{row_num}'].fill = calc_fill
        ws[f'E{row_num}'].border = border
        ws[f'E{row_num}'].alignment = center_align

        # Calls Offered (from forecast)
        ws[f'G{row_num}'] = row_data.get('Calls_Offered', 20)
        ws[f'G{row_num}'].border = border
        ws[f'G{row_num}'].alignment = center_align

        # AHT (from forecast)
        ws[f'H{row_num}'] = row_data.get('Average_Handle_Time_Seconds', 260)
        ws[f'H{row_num}'].border = border
        ws[f'H{row_num}'].alignment = center_align

        # Traffic Intensity: (Calls × AHT) / 900 seconds (15 min interval)
        ws[f'I{row_num}'] = f'=(G{row_num}*H{row_num})/900'
        ws[f'I{row_num}'].number_format = '0.00'
        ws[f'I{row_num}'].fill = calc_fill
        ws[f'I{row_num}'].border = border
        ws[f'I{row_num}'].alignment = center_align

        # Required Agents (from forecast - for comparison)
        ws[f'F{row_num}'] = row_data.get('Required_Agents', 8)
        ws[f'F{row_num}'].border = border
        ws[f'F{row_num}'].alignment = center_align

        # Erlang C Probability formula
        # This is a complex formula - using simplified approximation for demonstration
        # Full formula: P(W>0) = (A^N / N!) * (N/(N-A)) / [sum(A^k/k!) + (A^N/N!)*(N/(N-A))]
        # Simplified: Using Excel's SUMPRODUCT approach
        erlang_c_formula = f'''=IF(E{row_num}<=I{row_num},"Need More",
(POWER(E{row_num},I{row_num})/FACT(E{row_num})*(E{row_num}/(E{row_num}-I{row_num})))/
(SUMPRODUCT(POWER(I{row_num},ROW(INDIRECT("0:"&(E{row_num}-1))))/
FACT(ROW(INDIRECT("0:"&(E{row_num}-1)))))+
(POWER(E{row_num},I{row_num})/FACT(E{row_num})*(E{row_num}/(E{row_num}-I{row_num})))))'''

        ws[f'J{row_num}'] = erlang_c_formula
        ws[f'J{row_num}'].number_format = '0.0%'
        ws[f'J{row_num}'].fill = calc_fill
        ws[f'J{row_num}'].border = border
        ws[f'J{row_num}'].alignment = center_align

        # Service Level formula: 1 - (P(W>0) × e^(-(N-A)×(T/AHT)))
        # For 80/90 target: T = 90 seconds
        service_level_formula = f'''=IF(E{row_num}<=I{row_num},0,
IF(ISNUMBER(J{row_num}),1-(J{row_num}*EXP(-((E{row_num}-I{row_num})*(90/H{row_num})))),0))'''

        ws[f'K{row_num}'] = service_level_formula
        ws[f'K{row_num}'].number_format = '0.0%'
        ws[f'K{row_num}'].border = border
        ws[f'K{row_num}'].alignment = center_align

        # ASA (Average Speed of Answer): (P(W>0) × AHT) / (N - A)
        asa_formula = f'=IF(E{row_num}<=I{row_num},"Need More",IF(ISNUMBER(J{row_num}),(J{row_num}*H{row_num})/(E{row_num}-I{row_num}),0))'
        ws[f'L{row_num}'] = asa_formula
        ws[f'L{row_num}'].number_format = '0'
        ws[f'L{row_num}'].border = border
        ws[f'L{row_num}'].alignment = center_align

        # Occupancy: Traffic / Net Agents
        ws[f'M{row_num}'] = f'=IF(E{row_num}>0,I{row_num}/E{row_num},0)'
        ws[f'M{row_num}'].number_format = '0.0%'
        ws[f'M{row_num}'].border = border
        ws[f'M{row_num}'].alignment = center_align

        # Staffing Gap: Net Agents - Required Agents
        ws[f'N{row_num}'] = f'=E{row_num}-F{row_num}'
        ws[f'N{row_num}'].number_format = '0.0'
        ws[f'N{row_num}'].border = border
        ws[f'N{row_num}'].alignment = center_align

    last_data_row = data_start_row + min(len(forecast_df), 36) - 1

    # ===== CONDITIONAL FORMATTING =====
    # Highlight service levels below 80% in red
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(color="9C0006", bold=True)
    ws.conditional_formatting.add(
        f'K{data_start_row}:K{last_data_row}',
        CellIsRule(operator='lessThan', formula=['0.8'], fill=red_fill, font=red_font)
    )

    # Highlight occupancy > 85% in orange (too high)
    orange_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    ws.conditional_formatting.add(
        f'M{data_start_row}:M{last_data_row}',
        CellIsRule(operator='greaterThan', formula=['0.85'], fill=orange_fill)
    )

    # Highlight occupancy < 70% in blue (too low)
    blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    ws.conditional_formatting.add(
        f'M{data_start_row}:M{last_data_row}',
        CellIsRule(operator='lessThan', formula=['0.70'], fill=blue_fill)
    )

    # Highlight negative staffing gap (understaffed) in red
    ws.conditional_formatting.add(
        f'N{data_start_row}:N{last_data_row}',
        CellIsRule(operator='lessThan', formula=['0'], fill=red_fill, font=red_font)
    )

    # ===== SUMMARY DASHBOARD =====
    summary_row = last_data_row + 3

    ws.merge_cells(f'A{summary_row}:N{summary_row}')
    ws[f'A{summary_row}'] = 'SUMMARY DASHBOARD'
    ws[f'A{summary_row}'].fill = header_fill
    ws[f'A{summary_row}'].font = header_font
    ws[f'A{summary_row}'].alignment = center_align

    summary_row += 1

    # KPI: Average Service Level
    ws[f'A{summary_row}'] = 'Average Service Level:'
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'C{summary_row}'] = f'=AVERAGE(K{data_start_row}:K{last_data_row})'
    ws[f'C{summary_row}'].number_format = '0.0%'
    ws[f'C{summary_row}'].font = Font(bold=True, size=12)
    ws[f'C{summary_row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # KPI: Intervals Below Target
    ws[f'E{summary_row}'] = 'Intervals Below 80%:'
    ws[f'E{summary_row}'].font = Font(bold=True)
    ws[f'G{summary_row}'] = f'=COUNTIF(K{data_start_row}:K{last_data_row},"<0.8")'
    ws[f'G{summary_row}'].font = Font(bold=True, size=12, color="C00000")

    summary_row += 1

    # KPI: Average Occupancy
    ws[f'A{summary_row}'] = 'Average Occupancy:'
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'C{summary_row}'] = f'=AVERAGE(M{data_start_row}:M{last_data_row})'
    ws[f'C{summary_row}'].number_format = '0.0%'
    ws[f'C{summary_row}'].font = Font(bold=True, size=12)

    # KPI: Average ASA
    ws[f'E{summary_row}'] = 'Average ASA:'
    ws[f'E{summary_row}'].font = Font(bold=True)
    ws[f'G{summary_row}'] = f'=AVERAGE(L{data_start_row}:L{last_data_row})'
    ws[f'G{summary_row}'].number_format = '0" sec"'
    ws[f'G{summary_row}'].font = Font(bold=True, size=12)

    summary_row += 1

    # KPI: Total Staffing Gap
    ws[f'A{summary_row}'] = 'Total Staffing Gap:'
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'C{summary_row}'] = f'=SUM(N{data_start_row}:N{last_data_row})'
    ws[f'C{summary_row}'].number_format = '0.0'
    ws[f'C{summary_row}'].font = Font(bold=True, size=12)

    # Interpretation
    ws[f'E{summary_row}'] = '(Positive = Overstaffed, Negative = Understaffed)'
    ws[f'E{summary_row}'].font = Font(italic=True, size=9, color="7F7F7F")

    # Save the workbook
    output_filename = 'erlang_c_staffing_forecast.xlsx'
    workbook.save(output_filename)
    print(f"\n✓ Successfully created worksheet in '{output_filename}'")
    print(f"✓ Sample data populated for {last_data_row - data_start_row + 1} intervals")
    print("\nNext steps:")
    print("1. Open erlang_c_staffing_forecast.xlsx")
    print("2. Go to 'Schedule_Service_Level' worksheet")
    print("3. Adjust Shrinkage Rate in cell E4 (currently 25%)")
    print("4. Enter your Scheduled Agents in column D")
    print("5. Review Service Level %, ASA, Occupancy, and Staffing Gap outputs")
    print("\nConditional formatting applied:")
    print("  • Service Level < 80%: Highlighted in RED")
    print("  • Occupancy > 85%: Highlighted in ORANGE (overstaffed)")
    print("  • Occupancy < 70%: Highlighted in BLUE (understaffed)")
    print("  • Staffing Gap < 0: Highlighted in RED (need more agents)")

if __name__ == "__main__":
    print("=" * 70)
    print("SERVICE LEVEL CALCULATOR - Excel Worksheet Generator")
    print("=" * 70)
    print()

    create_service_level_worksheet()

    print("\n" + "=" * 70)
    print("COMPLETE!")
    print("=" * 70)
